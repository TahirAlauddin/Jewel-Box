import openpyxl
from datetime import datetime

def set_invoice_data(ws, invoice):
    """ Fill the invoice details """
    # Basic invoice information
    ws['H6'] = invoice.get('invoice_number', '')  # Invoice number
    ws['G7'] = invoice.get('date', datetime.today().strftime('%Y-%m-%d'))  # Date
    ws['B10'] = invoice.get('client', '')  # Client
    ws['F10'] = invoice.get('ship_to', '')  # Ship To
    ws['B10'] = invoice.get('to', '')  # Customer Address

def set_invoice_items(ws, items):
    """ Fill the invoice items """
    start_row = 16  # Starting row for items
    total_price = 0
    for i, item in enumerate(items):
         # Ensure that unit_price and quantity are neither None nor missing
        unit_price = item.get('unit_price') if item.get('unit_price') is not None else 0
        quantity = item.get('quantity') if item.get('quantity') is not None else 0

        # Convert unit_price and quantity to float and int respectively
        unit_price = float(unit_price)
        quantity = int(quantity)

        # Calculate total price
        price = unit_price * quantity

        row = start_row + i
        ws[f'A{row}'] = item.get('order_id', '')
        ws[f'B{row}'] = item.get('job_number', '')
        ws[f'C{row}'] = item.get('description', '')
        # D{row} is not missing, description is taking 2 columns
        ws[f'E{row}'] = item.get('type', '')
        ws[f'F{row}'] = item.get('quantity', '1')
        ws[f'G{row}'] = item.get('unit_price', '0')
        ws[f'H{row}'] = price  # Price calculation
        total_price += price
    ws[f'H33'] = total_price

def main(invoice, items, template_path, output_file):
    # Load the workbook and get the active worksheet
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Set invoice data and items
    set_invoice_data(ws, invoice)
    set_invoice_items(ws, items)


    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    invoice = {
        'invoice_number': 'INV1001',
        'date': '2024-05-22',
        'client': 'Acme Corp',
        'ship_to': '123 Market St',
        'to': '123 Market More...'
    }
    items = [
        {'order_id': 'ADL24001', 'job_number': 'JN1001', 'description': 'Gold Ring', 'type': 'Product', 'quantity': '2', 'unit_price': '150.00'},
        {'order_id': 'ADL24002', 'job_number': 'JN1002', 'description': 'Silver Necklace', 'type': 'Product', 'quantity': '1', 'unit_price': '100.00'}
    ]

    template_path = './Invoice.xlsx'
    output_file = 'invoice_generated.xlsx'
    
    main(invoice, items, template_path, output_file)
