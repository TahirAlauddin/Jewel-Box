import openpyxl
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage

def set_dummy_data(ws, data):
    """ Fill the cells in spreadsheet with given data """
    date = data.get('date', '')
    if date != '':
        date = date.date()
        date = date.strftime('%b %d, %Y')

    due_date = data.get('due_date', '')
    if due_date != '':
        due_date = due_date.strftime('%b %d, %Y')

    ws['B2'] = str(date)  # Date
    ws['D2'] = data.get('quantity', '')  # Quantity
    ws['D3'] = data.get('size', '')  # Size
    ws['D4'] = data.get('resize', '')  # Size
    ws['B3'] = str(due_date)  # Due date
    ws['B4'] = data.get('client', '')  # Client
    ws['D9'] = data.get('color', '')  # Resize
    ws['D5'] = data.get('type', '')  # Type
    ws['B5'] = data.get('order_number', '')  # Order #
    ws['B6'] = data.get('kt', '')  # KT
    ws['D6'] = data.get('ct', '')  # CT
    ws['B8'] = data.get('rush', '')  # Rush
    ws['D8'] = data.get('setter', '')  # Setter
    ws['B12'] = data.get('repair', '')  # Repair
    ws['B14'] = data.get('order', '')  # Order
    ws['B15'] = data.get('stamp', '')  # Stamp
    ws['B16'] = data.get('set', '')  # Set

    # List of StoneSpecification Objects
    stones = data.get('stones', [])

    for i in range(len(stones)):
        ws[f'A{26+i}'] = stones[i].stone_type  # Stone
        ws[f'B{26+i}'] = stones[i].cut  # Cut
        ws[f'C{26+i}'] = stones[i].id  # ID
        ws[f'D{26+i}'] = stones[i].quantity  # Quantity
        ws[f'E{26+i}'] = stones[i].length  # Length
        ws[f'F{26+i}'] = stones[i].width  # Width
        ws[f'G{26+i}'] = stones[i].height  # Height
        ws[f'H{26+i}'] = stones[i].carat_total  # CT

def adjust_row_heights(ws, text):
    for row in range(18, 24):
        if len(text) < 50:
            ws.row_dimensions[row].height = 10
        elif len(text) < 100:
            ws.row_dimensions[row].height = 12
        elif len(text) < 150:
            ws.row_dimensions[row].height = 14
        elif len(text) < 200:
            ws.row_dimensions[row].height = 15
        else:
            ws.row_dimensions[row].height = 15
            text = text[:200]

def insert_text(ws, text):
    ws['A18'] = text


def add_images(ws, barcode_image, jewelry_images):
    # Load the barcode image using Pillow
    barcode_img = Image.open(barcode_image)
    img = OpenpyxlImage(barcode_image)
    img.width = barcode_img.width / 2
    img.height = barcode_img.height / 2

    ws.add_image(img, 'E1')

    # Set initial position for placing the first jewelry image
    image_width = 220
    image_height = 140

    if isinstance(jewelry_images, list) and len(jewelry_images) >= 1:
        # Load and add the first jewelry image
        jewelry_img_path = jewelry_images[0]
        jewelry_img = OpenpyxlImage(jewelry_img_path)
        jewelry_img.width = image_width
        jewelry_img.height = image_height
        ws.add_image(jewelry_img, 'E2')

    if isinstance(jewelry_images, list) and len(jewelry_images) >= 2:
        # Calculate position for placing the second jewelry image below the first one
        jewelry_img_path = jewelry_images[1]
        jewelry_img = OpenpyxlImage(jewelry_img_path)
        jewelry_img.width = image_width
        jewelry_img.height = image_height
        ws.add_image(jewelry_img, 'E8')

    if isinstance(jewelry_images, str):
        jewelry_img_path = jewelry_images
        jewelry_img = OpenpyxlImage(jewelry_img_path)
        jewelry_img.width = image_width
        jewelry_img.height = image_height
        ws.add_image(jewelry_img, 'E2')


def main(data, text, barcode_image, jewelry_images, output_file):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.load_workbook('orders/Production Sheet.xlsx')
    ws = wb.active
    set_dummy_data(ws, data)
    adjust_row_heights(ws, text)
    insert_text(ws, text)
    add_images(ws, barcode_image, jewelry_images)

    # Save the workbook
    wb.save(output_file)

if __name__ == "__main__":
    from datetime import datetime, date
    
    data = {
        'date': datetime(2024, 9, 10),
        'quantity': '10',
        'size': '5',
        'due_date': date(2024, 8, 10),
        'client': 'John Doe',
        'resize': 'Yes',
        'order_number': 'AFK24001',
        'kt': '14K',
        'ct': '0.5',
        'rush': 'Yes',
        'setter': 'Smith',
        'type': 'Ring',
        'repair': 'Yes',
        'order': 'R001',
        'stamp': 'S925',
    }
    text = 'Check all stones before setting. And thats not it. Also make sure you are eating breakfast to stay healthy otherwise you will just get weaker. healthy otherwise you will just get weaker. healthy otherwise you will just get weaker'
    barcode_image = r'media\barcodes\PTK24002_-_barcode.png'
    jewelry_images = [r'C:\Users\Hi\OneDrive\Desktop\Projects\Jewel Box\jewelry.jpeg',
                    r'C:\Users\Hi\OneDrive\Desktop\Projects\Jewel Box\jewelry.jpeg']
    output_file = 'ascent_production_sheet.xlsx'

    main(data, text, barcode_image, jewelry_images, output_file)

